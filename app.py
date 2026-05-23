import os
import re
import sqlite3
from datetime import datetime, date, time, timedelta
from copy import copy
from unicodedata import normalize
from flask import Flask, render_template_string, request, redirect, url_for, send_file, send_from_directory, flash, session
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def resolver_data_dir():
    """
    Define una carpeta escribible para SQLite, fotos y exportaciones.

    En Render Free NO siempre se puede escribir en /var/data si no hay disco persistente.
    Por eso se prueba DATA_DIR y, si falla, se usa /tmp automáticamente.
    """
    candidatos = []

    env_dir = os.environ.get("DATA_DIR", "").strip()
    if env_dir:
        candidatos.append(env_dir)

    candidatos.extend([
        "/tmp",
        os.path.join(BASE_DIR, "data"),
        BASE_DIR,
    ])

    for carpeta in candidatos:
        try:
            os.makedirs(carpeta, exist_ok=True)
            prueba = os.path.join(carpeta, ".write_test")
            with open(prueba, "w", encoding="utf-8") as f:
                f.write("ok")
            os.remove(prueba)
            return carpeta
        except Exception as e:
            print(f"No se puede usar DATA_DIR={carpeta}: {e}")

    return "/tmp"


DATA_DIR = resolver_data_dir()

DB_PATH = os.path.join(DATA_DIR, "mbo.db")
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
EXPORT_DIR = os.path.join(DATA_DIR, "exports")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)

EXCEL_MBO = "HGP-SG3-OP-FR-033 Mantenimiento Básico Operacional (MBO)_May_26.xlsx"
SHEETS_MBO = [
    "PRESA",
    "COLCHON DE AIRE",
    "EDIFICIO CONTROL",
    "SE Paquillusi",
    "UG1",
    "UG2",
    "COMUNES",
]

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "mbo_chsg3_local_cambiar_en_produccion")


def limpiar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor).replace("\n", " ").replace("\r", " ").strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar(valor):
    texto = limpiar_texto(valor).upper()
    texto = normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto


def opciones_selector_valor(referencia):
    """
    Devuelve opciones tipo selector para referencias operativas.
    Si no reconoce una referencia especial, devuelve lista vacía y se usa input libre.
    """
    ref = normalizar(referencia)

    # OK/NOK
    if "OK" in ref and "NOK" in ref:
        return [
            {"value": "OK", "label": "OK"},
            {"value": "NOK", "label": "NOK"},
        ]

    # SI / NO
    # Cubre referencias como: SI/NO, SI / NO, SÍ / NO.
    if ("SI" in ref and "NO" in ref) or ("S/" in ref and "N/" in ref):
        return [
            {"value": "SI", "label": "SI"},
            {"value": "NO", "label": "NO"},
        ]

    # L: Local R: Remoto
    if "LOCAL" in ref and "REMOTO" in ref:
        return [
            {"value": "L", "label": "L"},
            {"value": "R", "label": "R"},
        ]

    # A: Auto / Automático  M: Manual
    # Cubre: A: Auto M: Manual, A: Automatico M: Manual, Automático/Manual.
    if (("AUTO" in ref or "AUTOMATICO" in ref) and "MANUAL" in ref):
        return [
            {"value": "A", "label": "A"},
            {"value": "M", "label": "M"},
        ]

    # C: Cerrado A: Abierto
    if "CERRADO" in ref and "ABIERTO" in ref:
        return [
            {"value": "C", "label": "C"},
            {"value": "A", "label": "A"},
        ]

    return []


def es_item_horas_operacion(item):
    """
    Detecta ítems de horas de operación.
    Se usa para convertir entradas tipo 20D15H20M10S a horas decimales.
    """
    unidad = normalizar(item.get("unidad", "") if isinstance(item, dict) else item["unidad"])
    descripcion = normalizar(item.get("descripcion", "") if isinstance(item, dict) else item["descripcion"])

    return (
        unidad in ("HRS", "HR", "H")
        or "HORAS DE OPERACION" in descripcion
        or "HORAS OPERACION" in descripcion
    )


def convertir_duracion_a_horas(valor):
    """
    Convierte formatos como:
    20D15H20M10S -> 495.3361
    15H20M10S    -> 15.3361
    20M10S       -> 0.3361

    Si el texto no tiene formato de duración, devuelve None.
    """
    texto_original = limpiar_texto(valor)
    if not texto_original:
        return None

    texto = normalizar(texto_original).replace(",", ".")
    texto = re.sub(r"\s+", "", texto)

    # Debe contener por lo menos una unidad D/H/M/S.
    if not re.search(r"[DHMS]", texto):
        return None

    patron = re.compile(
        r"^(?:(?P<d>\d+(?:\.\d+)?)D)?"
        r"(?:(?P<h>\d+(?:\.\d+)?)H)?"
        r"(?:(?P<m>\d+(?:\.\d+)?)M)?"
        r"(?:(?P<s>\d+(?:\.\d+)?)S)?$"
    )

    m = patron.match(texto)
    if not m:
        return None

    if not any(m.group(k) for k in ("d", "h", "m", "s")):
        return None

    dias = float(m.group("d") or 0)
    horas = float(m.group("h") or 0)
    minutos = float(m.group("m") or 0)
    segundos = float(m.group("s") or 0)

    total_horas = dias * 24 + horas + minutos / 60 + segundos / 3600

    # 4 decimales, quitando ceros sobrantes.
    return f"{total_horas:.4f}".rstrip("0").rstrip(".")


def buscar_excel_mbo():
    # Busca primero junto al app.py y luego en DATA_DIR.
    carpetas = []
    for carpeta in (BASE_DIR, DATA_DIR):
        if carpeta not in carpetas:
            carpetas.append(carpeta)

    for carpeta in carpetas:
        ruta = os.path.join(carpeta, EXCEL_MBO)
        if os.path.exists(ruta):
            return ruta

    for carpeta in carpetas:
        if not os.path.isdir(carpeta):
            continue
        for archivo in os.listdir(carpeta):
            nombre = archivo.upper()
            if nombre.endswith((".XLSX", ".XLSM")) and "MBO" in nombre and "FR-033" in nombre:
                return os.path.join(carpeta, archivo)

    return None


def conectar():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def crear_bd():
    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS items_mbo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            zona TEXT,
            nivel TEXT,
            sistema TEXT,
            equipo TEXT,
            descripcion TEXT,
            unidad TEXT,
            senal TEXT,
            seteos TEXT,
            referencia TEXT,
            fila_excel INTEGER,
            activo INTEGER DEFAULT 1
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS registros_mbo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            hora TEXT,
            usuario TEXT,
            zona TEXT,
            item_id INTEGER,
            nivel TEXT,
            sistema TEXT,
            equipo TEXT,
            descripcion TEXT,
            unidad TEXT,
            senal TEXT,
            seteos TEXT,
            referencia TEXT,
            valor TEXT,
            observacion TEXT,
            foto TEXT,
            fecha_registro TEXT,
            estado TEXT
        )
    """)

    conn.commit()
    conn.close()


def detectar_columnas(ws):
    """
    Detecta la fila de encabezados y las columnas principales del formato MBO.
    Funciona para PRESA, COLCHON DE AIRE, EDIFICIO CONTROL, SE Paquillusi, UG1, UG2 y COMUNES.
    """
    mejor = None

    for fila in range(1, min(ws.max_row, 20) + 1):
        columnas = {}
        for col in range(1, min(ws.max_column, 20) + 1):
            val = normalizar(ws.cell(fila, col).value)
            if not val:
                continue

            if "NIVEL" == val or val.startswith("NIVEL"):
                columnas["nivel"] = col
            elif "SISTEMA" == val or val.startswith("SISTEMA"):
                columnas["sistema"] = col
            elif "DESCRIPCION" in val:
                columnas["descripcion"] = col
            elif "UNIDAD" == val or val.startswith("UNIDAD"):
                columnas["unidad"] = col
            elif "SENAL" in val or "SEAL" in val:
                columnas["senal"] = col
            elif "SETEO" in val:
                columnas["seteos"] = col
            elif "REFERENCIA" in val:
                columnas["referencia"] = col

        if "descripcion" in columnas and "unidad" in columnas:
            mejor = {"fila_header": fila, **columnas}
            break

    if not mejor:
        return None

    # Equipo normalmente está justo antes de DESCRIPCIÓN, si existe una columna intermedia.
    desc_col = mejor.get("descripcion")
    if desc_col and desc_col > 1:
        posible_equipo = desc_col - 1
        usados = {v for k, v in mejor.items() if isinstance(v, int)}
        if posible_equipo not in usados:
            mejor["equipo"] = posible_equipo

    return mejor


def cargar_items_desde_excel(forzar=False):
    ruta_excel = buscar_excel_mbo()
    if not ruta_excel:
        return 0, "No se encontró el Excel MBO en la misma carpeta del app."

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) AS total FROM items_mbo")
    total_actual = cur.fetchone()["total"]

    if total_actual > 0 and not forzar:
        conn.close()
        return total_actual, "Los ítems ya estaban cargados."

    if forzar:
        cur.execute("DELETE FROM items_mbo")
        conn.commit()

    wb = load_workbook(ruta_excel, data_only=True)
    total_insertado = 0

    for nombre_hoja in SHEETS_MBO:
        if nombre_hoja not in wb.sheetnames:
            continue

        ws = wb[nombre_hoja]
        cols = detectar_columnas(ws)
        if not cols:
            continue

        fila_header = cols["fila_header"]
        ult_nivel = ""
        ult_sistema = ""
        ult_equipo = ""

        for fila in range(fila_header + 1, ws.max_row + 1):
            descripcion = limpiar_texto(ws.cell(fila, cols["descripcion"]).value)

            if not descripcion:
                continue

            desc_norm = normalizar(descripcion)
            if desc_norm in ("DESCRIPCION", "DESCRIPCIÓN"):
                continue
            if desc_norm.startswith("OBSERV") or desc_norm.startswith("FIRMA"):
                continue

            nivel = limpiar_texto(ws.cell(fila, cols.get("nivel", 0)).value) if cols.get("nivel") else ""
            sistema = limpiar_texto(ws.cell(fila, cols.get("sistema", 0)).value) if cols.get("sistema") else ""
            equipo = limpiar_texto(ws.cell(fila, cols.get("equipo", 0)).value) if cols.get("equipo") else ""

            if nivel:
                ult_nivel = nivel
            if sistema:
                ult_sistema = sistema
            if equipo:
                ult_equipo = equipo

            unidad = limpiar_texto(ws.cell(fila, cols.get("unidad", 0)).value) if cols.get("unidad") else ""
            senal = limpiar_texto(ws.cell(fila, cols.get("senal", 0)).value) if cols.get("senal") else ""
            seteos = limpiar_texto(ws.cell(fila, cols.get("seteos", 0)).value) if cols.get("seteos") else ""
            referencia = limpiar_texto(ws.cell(fila, cols.get("referencia", 0)).value) if cols.get("referencia") else ""

            cur.execute("""
                INSERT INTO items_mbo
                (zona, nivel, sistema, equipo, descripcion, unidad, senal, seteos, referencia, fila_excel, activo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            """, (
                nombre_hoja,
                ult_nivel,
                ult_sistema,
                ult_equipo,
                descripcion,
                unidad,
                senal,
                seteos,
                referencia,
                fila,
            ))
            total_insertado += 1

    conn.commit()
    conn.close()
    return total_insertado, f"Carga completada desde: {os.path.basename(ruta_excel)}"


def aplicar_ajustes_comunes():
    """
    Ajuste manual para la hoja COMUNES:
    - Ítems 19 al 29: BARRA A
    - Ítems 30 al 40: BARRA B

    Se actualiza el campo equipo para que la pantalla agrupe y diferencie correctamente
    ambas barras dentro de SALA ELÉCTRICA 400 V. También actualiza registros ya guardados.
    """
    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT id
        FROM items_mbo
        WHERE zona = ? AND activo = 1
        ORDER BY id
    """, ("COMUNES",))
    filas = cur.fetchall()

    for numero_item, fila in enumerate(filas, start=1):
        equipo_barra = ""
        if 19 <= numero_item <= 29:
            equipo_barra = "BARRA A"
        elif 30 <= numero_item <= 40:
            equipo_barra = "BARRA B"

        if equipo_barra:
            item_id = fila["id"]
            cur.execute("UPDATE items_mbo SET equipo = ? WHERE id = ?", (equipo_barra, item_id))
            cur.execute("""
                UPDATE registros_mbo
                SET equipo = ?
                WHERE zona = ? AND item_id = ?
            """, (equipo_barra, "COMUNES", item_id))

    conn.commit()
    conn.close()


def aplicar_ajustes_colchon_aire_solo_app():
    """
    Ajuste manual para COLCHON DE AIRE:
    el rótulo OPERADOR DE TURNO pertenece al bloque/firma del formato Excel,
    no es un punto de inspección. Por eso se oculta solo en el aplicativo.

    Importante: no se elimina ni se oculta la fila en el Excel exportado.
    """
    conn = conectar()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, descripcion, nivel, sistema, equipo, fila_excel
        FROM items_mbo
        WHERE zona = ?
        ORDER BY id
    """, ("COLCHON DE AIRE",))
    filas = cur.fetchall()

    for numero_item, fila in enumerate(filas, start=1):
        desc_norm = normalizar(fila["descripcion"])
        nivel_norm = normalizar(fila["nivel"])
        sistema_norm = normalizar(fila["sistema"])
        equipo_norm = normalizar(fila["equipo"])
        fila_excel = int(fila["fila_excel"] or 0)

        es_operador_turno = desc_norm == "OPERADOR DE TURNO"
        es_footer_accesos = nivel_norm == "ACCESOS" and es_operador_turno
        es_item_136_app = numero_item == 136 and es_operador_turno
        es_fila_footer_excel = fila_excel in (146, 136) and es_operador_turno

        if es_footer_accesos or es_item_136_app or es_fila_footer_excel:
            cur.execute("UPDATE items_mbo SET activo = 0 WHERE id = ?", (fila["id"],))

    conn.commit()
    conn.close()


def es_registro_operador_turno_colchon(r):
    """
    Evita escribir registros históricos del falso ítem OPERADOR DE TURNO
    dentro del cuerpo de datos. El rótulo debe quedar como footer del Excel.
    """
    try:
        zona = normalizar(r["zona"])
        descripcion = normalizar(r["descripcion"])
        nivel = normalizar(r["nivel"])
        fila_excel = int(r["fila_excel"] or 0)
    except Exception:
        return False

    return (
        zona == "COLCHON DE AIRE"
        and descripcion == "OPERADOR DE TURNO"
        and (nivel == "ACCESOS" or fila_excel in (146, 136))
    )


HTML_INDEX = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>MBO Diario CHSG3</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">

    <style>
        :root{
            --azul:#183b70;
            --azul2:#2457a6;
            --azul3:#e8f0fb;
            --gris:#f3f6fa;
            --borde:#d8e2ef;
            --texto:#111827;
            --muted:#64748b;
            --verde:#15803d;
            --verde2:#e8f8ee;
            --amarillo:#facc15;
            --rojo:#dc2626;
            --blanco:#ffffff;
            --shadow:0 10px 28px rgba(15,23,42,.10);
        }

        *{ box-sizing:border-box; }

        body{
            margin:0;
            font-family:Arial, sans-serif;
            background:
                radial-gradient(circle at top left, #dbeafe 0, transparent 32%),
                linear-gradient(180deg, #eef4fb 0%, #f8fafc 100%);
            color:var(--texto);
        }

        .wrap{
            max-width:1320px;
            margin:auto;
            padding:14px;
        }

        .top{
            background:linear-gradient(135deg, #163767, #244f92);
            color:white;
            padding:18px 20px;
            border-radius:20px;
            margin-bottom:14px;
            box-shadow:var(--shadow);
        }

        .top-row{
            display:flex;
            justify-content:space-between;
            gap:14px;
            align-items:center;
            flex-wrap:wrap;
        }

        .top h2{
            margin:0;
            font-size:24px;
            letter-spacing:.5px;
        }

        .top p{
            margin:6px 0 0 0;
            font-size:14px;
            opacity:.95;
        }

        .badge{
            background:rgba(255,255,255,.16);
            border:1px solid rgba(255,255,255,.28);
            padding:10px 14px;
            border-radius:999px;
            font-weight:bold;
            white-space:nowrap;
        }

        .menu{
            display:flex;
            gap:9px;
            flex-wrap:wrap;
            margin:14px 0;
            position:sticky;
            top:0;
            z-index:20;
            background:rgba(238,244,251,.88);
            backdrop-filter:blur(8px);
            padding:8px 0;
        }

        .menu a,
        .btn{
            background:var(--azul2);
            color:white;
            text-decoration:none;
            border:0;
            padding:11px 15px;
            border-radius:13px;
            font-weight:bold;
            cursor:pointer;
            box-shadow:0 4px 12px rgba(37,99,235,.18);
            transition:.15s ease;
            display:inline-flex;
            align-items:center;
            justify-content:center;
            gap:7px;
        }

        .menu a:hover,
        .btn:hover{
            transform:translateY(-1px);
            filter:brightness(1.05);
        }

        .menu a.sec{
            background:#4b5563;
        }

        .panel{
            background:rgba(255,255,255,.92);
            padding:14px;
            border-radius:20px;
            box-shadow:var(--shadow);
            margin-bottom:12px;
            border:1px solid rgba(216,226,239,.9);
        }

        label{
            font-weight:bold;
            display:block;
            margin-bottom:6px;
            color:#1f2937;
        }

        input,
        select,
        textarea{
            width:100%;
            box-sizing:border-box;
            padding:11px 12px;
            border:1px solid #cbd5e1;
            border-radius:12px;
            font-size:14px;
            background:white;
            outline:none;
            transition:.15s ease;
        }

        input:focus,
        select:focus,
        textarea:focus{
            border-color:#2563eb;
            box-shadow:0 0 0 3px rgba(37,99,235,.12);
        }

        .zona-nivel-grid{
            display:grid;
            grid-template-columns:minmax(280px, .9fr) minmax(420px, 1.6fr);
            gap:14px;
            align-items:start;
        }

        .nivel-card{
            min-width:0;
        }

        .nivel-scroll{
            display:flex;
            gap:8px;
            overflow-x:auto;
            padding:2px 2px 8px;
            scrollbar-width:thin;
            scroll-snap-type:x proximity;
        }

        .nivel-scroll::-webkit-scrollbar{
            height:8px;
        }

        .nivel-scroll::-webkit-scrollbar-thumb{
            background:#94a3b8;
            border-radius:999px;
        }

        .nivel-chip{
            flex:0 0 auto;
            min-width:105px;
            border:1px solid #c7d2fe;
            background:#eef2ff;
            color:#1e3a8a;
            padding:11px 14px;
            border-radius:999px;
            font-weight:bold;
            cursor:pointer;
            transition:.15s ease;
            scroll-snap-align:start;
        }

        .nivel-chip:hover{
            background:#dbeafe;
            transform:translateY(-1px);
        }

        .nivel-chip.active{
            background:linear-gradient(135deg, #173b70, #2457a6);
            color:white;
            border-color:#173b70;
            box-shadow:0 4px 12px rgba(30,64,175,.22);
        }

        .nivel-ayuda{
            font-size:12px;
            color:#64748b;
            margin:0 0 8px 0;
        }

        .grid{
            display:grid;
            grid-template-columns:repeat(4,1fr);
            gap:12px;
        }

        .ok{
            background:#dcfce7;
            color:#14532d;
            padding:10px 12px;
            border-radius:14px;
            margin:8px 0;
            border:1px solid #bbf7d0;
            font-weight:bold;
        }

        .warn{
            background:#fef9c3;
            color:#713f12;
            padding:10px 12px;
            border-radius:14px;
            margin:8px 0;
            border:1px solid #fde68a;
        }

        .small{
            font-size:12px;
            color:var(--muted);
            margin-bottom:0;
        }

        .toolbar{
            display:grid;
            grid-template-columns:1.2fr .8fr .8fr .8fr;
            gap:10px;
            align-items:end;
        }

        .toolbar .mini-btn{
            background:#eef2ff;
            color:#1e3a8a;
            border:1px solid #c7d2fe;
            padding:11px;
            border-radius:12px;
            font-weight:bold;
            cursor:pointer;
            transition:.15s ease;
        }

        .toolbar .mini-btn:hover{
            background:#dbeafe;
        }

        .progress-card{
            display:grid;
            grid-template-columns:1fr auto;
            gap:12px;
            align-items:center;
            margin-top:12px;
            padding:12px;
            background:var(--azul3);
            border-radius:16px;
            border:1px solid #bfdbfe;
        }

        .progress-bar{
            height:12px;
            background:white;
            border-radius:999px;
            overflow:hidden;
            border:1px solid #dbeafe;
        }

        .progress-fill{
            height:100%;
            width:0%;
            background:linear-gradient(90deg, #22c55e, #15803d);
            transition:.2s ease;
        }

        .progress-text{
            font-weight:bold;
            color:#1e3a8a;
            white-space:nowrap;
        }

        .tablebox{
            background:white;
            overflow:auto;
            border-radius:18px;
            box-shadow:var(--shadow);
            border:1px solid var(--borde);
            max-height:68vh;
        }

        table{
            width:100%;
            border-collapse:separate;
            border-spacing:0;
            font-size:13px;
        }

        th{
            background:#dbe8f7;
            color:#111827;
            padding:10px 8px;
            border-bottom:1px solid #334155;
            border-right:1px solid #94a3b8;
            text-align:center;
            position:sticky;
            top:0;
            z-index:5;
            white-space:nowrap;
        }

        td{
            border-bottom:1px solid #d7e0ea;
            border-right:1px solid #e2e8f0;
            padding:7px;
            vertical-align:middle;
            background:#fff;
            transition:.12s ease;
        }

        tr:nth-child(even).item-row td{
            background:#f3f8ff;
        }

        tr.item-row:hover td{
            background:#fff7d6 !important;
        }

        tr.item-row.row-complete td{
            background:var(--verde2) !important;
        }

        .grupo td{
            background:linear-gradient(90deg, #173b70, #2457a6) !important;
            color:white;
            font-weight:bold;
            text-align:left;
            position:sticky;
            top:39px;
            z-index:4;
            cursor:pointer;
            padding:9px 10px;
        }

        .grupo td::before{
            content:"▾ ";
            font-weight:bold;
        }

        .grupo.collapsed td::before{
            content:"▸ ";
        }

        .desc{
            min-width:280px;
            font-weight:600;
        }

        .center{
            text-align:center;
        }

        .ref{
            min-width:155px;
            font-size:12px;
            white-space:pre-wrap;
            color:#374151;
        }

        .sis{
            min-width:190px;
        }

        input.valor{
            min-width:95px;
            text-align:center;
            font-weight:bold;
        }

        .choice-group{
            display:flex;
            flex-wrap:wrap;
            gap:8px;
            align-items:center;
            justify-content:center;
            min-width:150px;
        }

        .choice-pill{
            display:inline-flex;
            margin:0;
            cursor:pointer;
        }

        .choice-pill input{
            display:none;
        }

        .choice-pill span{
            display:inline-flex;
            align-items:center;
            justify-content:center;
            min-width:58px;
            padding:10px 12px;
            border-radius:999px;
            border:1px solid #c7d2fe;
            background:#eef2ff;
            color:#1e3a8a;
            font-weight:bold;
            font-size:14px;
            transition:.15s ease;
            user-select:none;
        }

        .choice-pill input:checked + span{
            background:linear-gradient(135deg, #15803d, #2f8f46);
            color:white;
            border-color:#15803d;
            box-shadow:0 4px 12px rgba(21,128,61,.22);
        }

        .choice-pill:hover span{
            transform:translateY(-1px);
            background:#dbeafe;
        }

        .choice-pill input:checked + span:hover{
            background:linear-gradient(135deg, #15803d, #2f8f46);
        }

        textarea.obs{
            min-width:160px;
            resize:vertical;
        }

        input.foto{
            min-width:145px;
            font-size:12px;
            padding:8px;
        }

        .preview{
            margin-top:5px;
            width:44px;
            height:44px;
            object-fit:cover;
            border-radius:10px;
            border:1px solid #cbd5e1;
            display:none;
        }

        .sticky{
            position:sticky;
            bottom:0;
            background:rgba(238,244,251,.92);
            backdrop-filter:blur(8px);
            padding:10px 0;
            z-index:10;
        }

        .save{
            width:100%;
            background:linear-gradient(135deg, #15803d, #2f8f46);
            font-size:18px;
            padding:16px;
            border-radius:16px;
        }

        .hidden-by-filter{
            display:none !important;
        }

        .pill{
            display:inline-block;
            padding:4px 8px;
            border-radius:999px;
            background:#eef2ff;
            color:#1e3a8a;
            font-size:12px;
            font-weight:bold;
        }

        @media(max-width:1000px){
            .grid,
            .toolbar,
            .zona-nivel-grid{
                grid-template-columns:1fr 1fr;
            }

            table{
                font-size:12px;
            }

            th, td{
                padding:6px;
            }
        }

        /* FORMATO CELULAR CORREGIDO */
        @media(max-width:650px){
            body{
                background:#eef4fb;
            }

            .wrap{
                padding:8px;
                max-width:100%;
            }

            .top{
                border-radius:16px;
                padding:15px;
            }

            .top h2{
                font-size:20px;
            }

            .badge{
                width:100%;
                text-align:center;
            }

            .menu{
                position:relative;
                display:grid;
                grid-template-columns:1fr;
                gap:8px;
            }

            .menu a,
            .btn{
                width:100%;
                min-height:44px;
            }

            .grid,
            .toolbar,
            .zona-nivel-grid{
                grid-template-columns:1fr;
            }

            .nivel-scroll{
                padding-bottom:10px;
            }

            .nivel-chip{
                min-width:92px;
                padding:10px 12px;
                font-size:13px;
            }

            .progress-card{
                grid-template-columns:1fr;
            }

            .tablebox{
                max-height:none;
                overflow:visible;
                background:transparent;
                border:0;
                box-shadow:none;
                border-radius:0;
            }

            table,
            thead,
            tbody,
            th,
            td,
            tr{
                display:block;
                width:100%;
            }

            thead{
                display:none;
            }

            .grupo{
                margin:12px 0 8px;
            }

            .grupo td{
                position:relative;
                top:auto;
                border-radius:14px;
                margin:0;
                padding:11px 12px;
                font-size:13px;
                line-height:1.35;
                border:0;
                box-shadow:0 5px 14px rgba(15,23,42,.10);
            }

            tr.item-row{
                background:white;
                margin:10px 0 14px;
                border-radius:18px;
                box-shadow:0 8px 22px rgba(15,23,42,.10);
                overflow:hidden;
                border:1px solid var(--borde);
            }

            tr.item-row td{
                border:0;
                border-bottom:1px solid #eef2f7;
                background:white !important;
                padding:9px 12px;
                display:grid;
                grid-template-columns:96px minmax(0, 1fr);
                column-gap:10px;
                align-items:start;
                font-size:13px;
                line-height:1.35;
                text-align:left !important;
                overflow-wrap:anywhere;
                word-break:normal;
            }

            tr.item-row.row-complete td{
                background:var(--verde2) !important;
            }

            tr.item-row td::before{
                content:attr(data-label);
                font-weight:bold;
                color:#334155;
                font-size:12px;
                line-height:1.25;
                padding-top:2px;
            }

            tr.item-row td.mobile-head{
                display:flex;
                justify-content:space-between;
                align-items:center;
                gap:10px;
                background:#f8fbff !important;
                border-bottom:1px solid #dbe8f7;
            }

            tr.item-row.row-complete td.mobile-head{
                background:#e0f6e7 !important;
            }

            tr.item-row td.mobile-head::before{
                content:"";
                display:none;
            }

            .mobile-title{
                font-weight:bold;
                color:#111827;
                font-size:14px;
                line-height:1.35;
            }

            .mobile-num{
                flex:0 0 auto;
            }

            .desc-cell{
                display:block !important;
            }

            .desc-cell::before{
                content:"DESCRIPCIÓN";
                display:block;
                margin-bottom:5px;
                font-size:11px;
                letter-spacing:.04em;
                color:#64748b;
            }

            .desc{
                min-width:0;
                font-size:14px;
            }

            .sis,
            .ref,
            textarea.obs,
            input.valor,
            input.foto{
                min-width:0;
                width:100%;
            }

            .ref{
                white-space:normal;
            }

            input.valor,
            textarea.obs,
            input.foto{
                font-size:15px;
                padding:12px;
            }

            .choice-group{
                justify-content:flex-start;
                min-width:0;
                width:100%;
            }

            .choice-pill span{
                min-width:56px;
                padding:10px 13px;
                font-size:14px;
            }

            textarea.obs{
                min-height:44px;
            }

            .preview{
                width:70px;
                height:70px;
            }

            .sticky{
                position:sticky;
                bottom:0;
                padding:8px 0;
                background:rgba(238,244,251,.96);
            }

            .save{
                font-size:16px;
                min-height:54px;
            }
        }
    </style>
</head>

<body>
<div class="wrap">

    <div class="top">
        <div class="top-row">
            <div>
                <h2>MBO DIARIO CHSG3</h2>
                <p>Registro ordenado por hoja / sistema / equipo</p>
            </div>
            <div class="badge">Ítems cargados: {{ total_items }}</div>
        </div>
    </div>

    {% with messages = get_flashed_messages() %}
      {% if messages %}
        {% for m in messages %}<div class="ok">{{ m }}</div>{% endfor %}
      {% endif %}
    {% endwith %}

    <div class="menu">
        <a href="{{ url_for('index') }}">Registrar</a>
        <a class="sec" href="{{ url_for('registros') }}">Ver registros</a>
        <a class="sec" href="{{ url_for('exportar_excel', zona=zona) }}">Exportar esta hoja</a>
        <a class="sec" href="{{ url_for('exportar_excel') }}">Exportar todas las hojas</a>
        <a class="sec" href="{{ url_for('recargar_items') }}" onclick="return confirm('¿Recargar ítems desde el Excel? No borra registros, solo actualiza la lista de ítems.')">Recargar ítems</a>
        <a class="sec" href="{{ url_for('logout') }}">Salir</a>
    </div>

    {% if total_items == 0 %}
        <div class="warn">
            No hay ítems cargados. Verifica que el Excel MBO esté en la misma carpeta que este aplicativo.
        </div>
    {% endif %}

    <div class="panel">
        <div class="zona-nivel-grid">
            <div>
                <form method="GET" action="{{ url_for('index') }}">
                    <label>Hoja / Zona</label>
                    <select name="zona" onchange="this.form.submit()">
                        {% for z in zonas %}
                            <option value="{{ z }}" {% if z == zona %}selected{% endif %}>{{ z }}</option>
                        {% endfor %}
                    </select>
                </form>
            </div>

            <div class="nivel-card">
                <label>Filtro rápido por nivel / acceso</label>
                <p class="nivel-ayuda">Desliza la lista y toca un nivel para mostrar solo esos ítems.</p>
                <div class="nivel-scroll" id="nivelScroll">
                    <button type="button" class="nivel-chip active" data-nivel="">Todos</button>
                    {% for n in niveles %}
                        <button type="button" class="nivel-chip" data-nivel="{{ n }}">{{ n }}</button>
                    {% endfor %}
                </div>
            </div>
        </div>
    </div>

    <form method="POST" action="{{ url_for('guardar_zona') }}" enctype="multipart/form-data">
        <input type="hidden" name="zona" value="{{ zona }}">

        <div class="panel">
            <div class="grid">
                <div>
                    <label>Fecha</label>
                    <input type="date" name="fecha" id="fechaRegistro" value="{{ fecha_hoy }}" required>
                </div>
                <div>
                    <label>Hora</label>
                    <input type="time" name="hora" id="horaRegistro" value="{{ hora_actual }}" required>
                </div>
                <div>
                    <label>Operador</label>
                    <input type="text" name="usuario" id="operadorRegistro" value="{{ ultimo_operador }}" placeholder="Nombre del operador" required autocomplete="name">
                </div>
                <div>
                    <label>Hoja seleccionada</label>
                    <input type="text" value="{{ zona }}" readonly>
                </div>
            </div>

            <p class="small">Solo se mostrarán por defecto los ítems pendientes. Los llenados se ocultan para evitar confusión.</p>
        </div>

        {% if items %}

        <div class="panel">
            <div class="toolbar">
                <div>
                    <label>Buscar ítem</label>
                    <input type="text" id="buscarItem" placeholder="Buscar por descripción, sistema, equipo, nivel, unidad...">
                </div>

                <button type="button" class="mini-btn" id="btnTodos">Mostrar todos</button>
                <button type="button" class="mini-btn" id="btnPendientes">Solo pendientes</button>
                <button type="button" class="mini-btn" id="btnLlenados">Solo llenados</button>
            </div>

            <div class="progress-card">
                <div>
                    <div class="progress-bar">
                        <div class="progress-fill" id="progressFill"></div>
                    </div>
                </div>
                <div class="progress-text" id="progressText">0 / {{ items|length }} llenados</div>
            </div>
        </div>

        <div class="tablebox">
            <table id="tablaMBO">
                <thead>
                    <tr>
                        <th>N°</th>
                        <th>NIVEL</th>
                        <th>SISTEMA / EQUIPO</th>
                        <th>DESCRIPCIÓN</th>
                        <th>UNIDAD</th>
                        <th>SEÑAL</th>
                        <th>REFERENCIA</th>
                        <th>VALOR</th>
                        <th>OBSERVACIÓN</th>
                        <th>FOTO</th>
                    </tr>
                </thead>

                <tbody>
                {% set ns = namespace(grupo='') %}

                {% for item in items %}
                    {% set grupo_actual = (item['nivel'] or '-') ~ ' | ' ~ (item['sistema'] or '-') ~ ' | ' ~ (item['equipo'] or '-') %}

                    {% if grupo_actual != ns.grupo %}
                        <tr class="grupo group-row" data-nivel="{{ item['nivel'] or '' }}">
                            <td colspan="10">
                                NIVEL: {{ item['nivel'] or '-' }}
                                &nbsp;&nbsp; | &nbsp;&nbsp;
                                SISTEMA: {{ item['sistema'] or '-' }}
                                {% if item['equipo'] %}
                                    &nbsp;&nbsp; | &nbsp;&nbsp; EQUIPO: {{ item['equipo'] }}
                                {% endif %}
                            </td>
                        </tr>
                        {% set ns.grupo = grupo_actual %}
                    {% endif %}

                    <tr class="item-row"
                        data-nivel="{{ item['nivel'] or '' }}"
                        data-search="{{ (item['nivel'] ~ ' ' ~ item['sistema'] ~ ' ' ~ item['equipo'] ~ ' ' ~ item['descripcion'] ~ ' ' ~ item['unidad'] ~ ' ' ~ item['senal'] ~ ' ' ~ item['referencia'])|lower }}">

                        <td class="mobile-head" data-label="N°">
                            <span class="pill mobile-num">{{ loop.index }}</span>
                            <span class="mobile-title">{{ item['descripcion'] }}</span>
                        </td>

                        <td class="center" data-label="NIVEL">{{ item['nivel'] or '' }}</td>

                        <td class="sis" data-label="SISTEMA">
                            {{ item['sistema'] or '' }}
                            {% if item['equipo'] %}
                                <br><b>{{ item['equipo'] }}</b>
                            {% endif %}
                        </td>

                        <td class="desc desc-cell" data-label="DESCRIPCIÓN">{{ item['descripcion'] }}</td>

                        <td class="center" data-label="UNIDAD">{{ item['unidad'] or '' }}</td>

                        <td class="center" data-label="SEÑAL">{{ item['senal'] or '' }}</td>

                        <td class="ref" data-label="REFERENCIA">{{ item['referencia'] or '' }}</td>

                        <td data-label="VALOR">
                            {% if item['valor_opciones'] %}
                                <div class="choice-group">
                                    {% for op in item['valor_opciones'] %}
                                        <label class="choice-pill">
                                            <input class="campo-control valor-choice" type="radio" name="valor_{{ item['id'] }}" value="{{ op['value'] }}">
                                            <span>{{ op['label'] }}</span>
                                        </label>
                                    {% endfor %}
                                </div>
                            {% else %}
                                <input class="valor campo-control" type="text" name="valor_{{ item['id'] }}" placeholder="{% if item.get('convertir_horas') %}Ej: 20D15H20M10S{% else %}Valor{% endif %}" data-convert-horas="{% if item.get('convertir_horas') %}1{% else %}0{% endif %}">
                            {% endif %}
                        </td>

                        <td data-label="OBSERVACIÓN">
                            <textarea class="obs campo-control" name="obs_{{ item['id'] }}" rows="1" placeholder="Observación"></textarea>
                        </td>

                        <td data-label="FOTO">
                            <input class="foto campo-control" type="file" name="foto_{{ item['id'] }}" accept="image/*" capture="environment">
                            <img class="preview" alt="Vista previa">
                        </td>
                    </tr>
                {% endfor %}
                </tbody>
            </table>
        </div>

        <div class="sticky">
            <button class="btn save" type="submit">Guardar registros de {{ zona }}</button>
        </div>

        {% endif %}
    </form>
</div>

<script>
document.addEventListener("DOMContentLoaded", function(){
    const rows = Array.from(document.querySelectorAll(".item-row"));
    const groups = Array.from(document.querySelectorAll(".group-row"));
    const searchInput = document.getElementById("buscarItem");
    const progressFill = document.getElementById("progressFill");
    const progressText = document.getElementById("progressText");
    const nivelChips = Array.from(document.querySelectorAll(".nivel-chip"));

    let filtroModo = "pendientes";
    let filtroNivel = "";

    const formMbo = document.querySelector("form");
    const fechaRegistro = document.getElementById("fechaRegistro");
    const horaRegistro = document.getElementById("horaRegistro");
    const operadorRegistro = document.getElementById("operadorRegistro");

    function pad2(n){
        return String(n).padStart(2, "0");
    }

    function aplicarFechaHoraActual(){
        const ahora = new Date();

        if(fechaRegistro){
            const yyyy = ahora.getFullYear();
            const mm = pad2(ahora.getMonth() + 1);
            const dd = pad2(ahora.getDate());
            fechaRegistro.value = `${yyyy}-${mm}-${dd}`;
        }

        if(horaRegistro){
            horaRegistro.value = `${pad2(ahora.getHours())}:${pad2(ahora.getMinutes())}`;
        }
    }

    // Fecha y hora por defecto del momento real de registro.
    aplicarFechaHoraActual();
    setInterval(aplicarFechaHoraActual, 30000);

    // Recordar el último operador en este celular/PC.
    if(operadorRegistro){
        const operadorGuardado = localStorage.getItem("mbo_ultimo_operador") || "";
        if(!operadorRegistro.value.trim() && operadorGuardado){
            operadorRegistro.value = operadorGuardado;
        }

        operadorRegistro.addEventListener("input", function(){
            localStorage.setItem("mbo_ultimo_operador", operadorRegistro.value.trim());
        });
    }

    if(formMbo){
        formMbo.addEventListener("submit", function(){
            // Justo antes de guardar se toma la hora actual del celular/PC.
            aplicarFechaHoraActual();

            if(operadorRegistro && operadorRegistro.value.trim()){
                localStorage.setItem("mbo_ultimo_operador", operadorRegistro.value.trim());
            }
        });
    }


    function convertirDuracionHoras(texto){
        let raw = (texto || "").trim().toUpperCase().replace(",", ".");
        raw = raw.normalize("NFD").replace(/[\u0300-\u036f]/g, "");
        raw = raw.replace(/\s+/g, "");

        if(!/[DHMS]/.test(raw)) return null;

        const patron = /^(?:(\d+(?:\.\d+)?)D)?(?:(\d+(?:\.\d+)?)H)?(?:(\d+(?:\.\d+)?)M)?(?:(\d+(?:\.\d+)?)S)?$/;
        const m = raw.match(patron);
        if(!m) return null;

        const d = parseFloat(m[1] || "0");
        const h = parseFloat(m[2] || "0");
        const min = parseFloat(m[3] || "0");
        const s = parseFloat(m[4] || "0");

        if(d === 0 && h === 0 && min === 0 && s === 0) return null;

        const total = d * 24 + h + min / 60 + s / 3600;
        return parseFloat(total.toFixed(4)).toString();
    }

    function aplicarConversionHoras(campo){
        if(!campo || campo.dataset.convertHoras !== "1") return;

        const convertido = convertirDuracionHoras(campo.value);
        if(convertido !== null){
            campo.value = convertido;
            campo.title = "Convertido automáticamente a horas";
        }
    }

    function rowTieneDato(row){
        const valor = row.querySelector("input.valor");
        const obs = row.querySelector("textarea.obs");
        const foto = row.querySelector("input.foto");

        const valorChoice = row.querySelector("input.valor-choice:checked");

        const tieneValor = (valor && valor.value.trim() !== "") || !!valorChoice;
        const tieneObs = obs && obs.value.trim() !== "";
        const tieneFoto = foto && foto.files && foto.files.length > 0;

        return tieneValor || tieneObs || tieneFoto;
    }

    function actualizarCompletados(){
        let completos = 0;

        rows.forEach(row => {
            const lleno = rowTieneDato(row);
            row.classList.toggle("row-complete", lleno);
            if(lleno) completos++;
        });

        const total = rows.length || 1;
        const pct = Math.round((completos / total) * 100);

        if(progressFill){
            progressFill.style.width = pct + "%";
        }

        if(progressText){
            progressText.textContent = completos + " / " + rows.length + " llenados";
        }
    }

    function actualizarVisibilidadGrupos(){
        groups.forEach(group => {
            let next = group.nextElementSibling;
            let visible = false;

            while(next && !next.classList.contains("group-row")){
                if(next.classList.contains("item-row") && !next.classList.contains("hidden-by-filter") && next.style.display !== "none"){
                    visible = true;
                    break;
                }
                next = next.nextElementSibling;
            }

            group.classList.toggle("hidden-by-filter", !visible);
        });
    }

    function aplicarFiltro(){
        const texto = (searchInput ? searchInput.value.trim().toLowerCase() : "");

        rows.forEach(row => {
            const lleno = rowTieneDato(row);
            const coincideTexto = row.dataset.search.includes(texto);
            const coincideNivel = !filtroNivel || row.dataset.nivel === filtroNivel;

            let coincideModo = true;
            if(filtroModo === "pendientes") coincideModo = !lleno;
            if(filtroModo === "llenados") coincideModo = lleno;

            row.classList.toggle("hidden-by-filter", !(coincideTexto && coincideModo && coincideNivel));
        });

        actualizarVisibilidadGrupos();
    }

    document.querySelectorAll(".campo-control").forEach(campo => {
        campo.addEventListener("input", function(){
            actualizarCompletados();
            aplicarFiltro();
        });

        campo.addEventListener("blur", function(){
            if(campo.classList.contains("valor")){
                aplicarConversionHoras(campo);
                actualizarCompletados();
                aplicarFiltro();
            }
        });

        campo.addEventListener("change", function(){
            if(campo.classList.contains("valor")){
                aplicarConversionHoras(campo);
            }

            actualizarCompletados();
            aplicarFiltro();

            if(campo.type === "file"){
                const img = campo.parentElement.querySelector(".preview");

                if(img && campo.files && campo.files[0]){
                    img.src = URL.createObjectURL(campo.files[0]);
                    img.style.display = "block";
                }
            }
        });
    });

    if(searchInput){
        searchInput.addEventListener("input", aplicarFiltro);
    }

    nivelChips.forEach(chip => {
        chip.addEventListener("click", function(){
            filtroNivel = chip.dataset.nivel || "";

            nivelChips.forEach(c => c.classList.remove("active"));
            chip.classList.add("active");

            aplicarFiltro();
        });
    });

    const btnTodos = document.getElementById("btnTodos");
    const btnPendientes = document.getElementById("btnPendientes");
    const btnLlenados = document.getElementById("btnLlenados");

    if(btnTodos){
        btnTodos.addEventListener("click", function(){
            filtroModo = "todos";
            filtroNivel = "";
            if(searchInput) searchInput.value = "";

            nivelChips.forEach(c => c.classList.remove("active"));
            const chipTodos = document.querySelector(".nivel-chip[data-nivel='']");
            if(chipTodos) chipTodos.classList.add("active");

            rows.forEach(row => row.style.display = "");
            groups.forEach(group => group.classList.remove("collapsed"));

            aplicarFiltro();
        });
    }

    if(btnPendientes){
        btnPendientes.addEventListener("click", function(){
            filtroModo = "pendientes";
            aplicarFiltro();
        });
    }

    if(btnLlenados){
        btnLlenados.addEventListener("click", function(){
            filtroModo = "llenados";
            aplicarFiltro();
        });
    }

    groups.forEach(group => {
        group.addEventListener("click", function(){
            group.classList.toggle("collapsed");

            let next = group.nextElementSibling;
            const colapsar = group.classList.contains("collapsed");

            while(next && !next.classList.contains("group-row")){
                if(next.classList.contains("item-row")){
                    next.style.display = colapsar ? "none" : "";
                }
                next = next.nextElementSibling;
            }

            actualizarVisibilidadGrupos();
        });
    });

    actualizarCompletados();
    aplicarFiltro();
});
</script>

</body>
</html>
"""

HTML_REGISTROS = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Registros MBO</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body { font-family: Arial, sans-serif; background:#eef2f6; margin:0; padding:12px; }
        .wrap { max-width:1100px; margin:auto; }
        .top { background:#003b6f; color:white; padding:16px; border-radius:14px; margin-bottom:12px; }
        .menu { display:flex; gap:8px; flex-wrap:wrap; margin:12px 0; }
        .menu a, button { background:#005baa; color:white; text-decoration:none; border:0; padding:10px 13px; border-radius:10px; font-weight:bold; cursor:pointer; }
        .menu a.sec { background:#4b5563; }
        .panel { background:white; padding:14px; border-radius:14px; box-shadow:0 2px 8px rgba(0,0,0,.10); margin-bottom:12px; }
        .grid { display:grid; grid-template-columns:repeat(4,1fr); gap:10px; }
        label { font-weight:bold; display:block; margin-top:10px; }
        input, select { width:100%; box-sizing:border-box; padding:10px; border:1px solid #cbd5e1; border-radius:10px; font-size:15px; margin-top:4px; }
        table { width:100%; border-collapse:collapse; font-size:13px; background:white; }
        th { background:#005baa; color:white; padding:8px; position:sticky; top:0; }
        td { border:1px solid #d9e2ec; padding:7px; vertical-align:top; }
        .tablebox { overflow:auto; max-height:70vh; border-radius:12px; }
        .foto { color:#005baa; font-weight:bold; }
        @media(max-width:700px){ .grid { grid-template-columns:1fr; } table{font-size:11px;} }
    </style>
</head>
<body>
<div class="wrap">
    <div class="top">
        <h2>REGISTROS MBO</h2>
        <p>Total mostrado: <b>{{ registros|length }}</b></p>
    </div>

    <div class="menu">
        <a href="{{ url_for('index') }}">Registrar</a>
        <a class="sec" href="{{ url_for('exportar_excel', fecha=fecha, zona=zona) }}">Exportar Excel filtrado</a>
        <a class="sec" href="{{ url_for('logout') }}">Salir</a>
    </div>

    <div class="panel">
        <form method="GET" action="{{ url_for('registros') }}">
            <div class="grid">
                <div>
                    <label>Fecha</label>
                    <input type="date" name="fecha" value="{{ fecha }}">
                </div>
                <div>
                    <label>Zona</label>
                    <select name="zona">
                        <option value="">Todas</option>
                        {% for z in zonas %}
                            <option value="{{ z }}" {% if z == zona %}selected{% endif %}>{{ z }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div style="align-self:end;">
                    <button type="submit">Filtrar</button>
                </div>
            </div>
        </form>
    </div>

    <div class="tablebox">
        <table>
            <thead>
                <tr>
                    <th>ID</th><th>Fecha</th><th>Hora</th><th>Operador</th><th>Zona</th>
                    <th>Sistema</th><th>Equipo</th><th>Descripción</th><th>Valor</th>
                    <th>Unidad</th><th>Obs.</th><th>Foto</th>
                </tr>
            </thead>
            <tbody>
                {% for r in registros %}
                <tr>
                    <td>{{ r['id'] }}</td>
                    <td>{{ r['fecha'] }}</td>
                    <td>{{ r['hora'] }}</td>
                    <td>{{ r['usuario'] }}</td>
                    <td>{{ r['zona'] }}</td>
                    <td>{{ r['sistema'] }}</td>
                    <td>{{ r['equipo'] }}</td>
                    <td>{{ r['descripcion'] }}</td>
                    <td>{{ r['valor'] }}</td>
                    <td>{{ r['unidad'] }}</td>
                    <td>{{ r['observacion'] }}</td>
                    <td>
                        {% if r['foto'] %}
                            <a class="foto" href="{{ url_for('ver_foto', filename=r['foto']) }}" target="_blank">Ver</a>
                        {% endif %}
                    </td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</div>
</body>
</html>
"""


HTML_LOGIN = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Ingreso MBO CHSG3</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body{
            margin:0;
            min-height:100vh;
            display:flex;
            align-items:center;
            justify-content:center;
            font-family:Arial, sans-serif;
            background:linear-gradient(135deg,#e8f0fb,#f8fafc);
            color:#111827;
            padding:18px;
        }
        .card{
            width:100%;
            max-width:420px;
            background:white;
            border-radius:22px;
            padding:24px;
            box-shadow:0 16px 36px rgba(15,23,42,.14);
            border:1px solid #d8e2ef;
        }
        .top{
            background:linear-gradient(135deg,#163767,#244f92);
            color:white;
            padding:18px;
            border-radius:18px;
            margin-bottom:18px;
        }
        h2{ margin:0; font-size:24px; }
        p{ color:#64748b; }
        label{ font-weight:bold; display:block; margin-bottom:7px; }
        input{
            width:100%;
            box-sizing:border-box;
            padding:13px;
            border:1px solid #cbd5e1;
            border-radius:13px;
            font-size:16px;
            outline:none;
        }
        input:focus{
            border-color:#2563eb;
            box-shadow:0 0 0 3px rgba(37,99,235,.12);
        }
        button{
            width:100%;
            margin-top:14px;
            padding:14px;
            border:0;
            border-radius:14px;
            background:#2457a6;
            color:white;
            font-weight:bold;
            font-size:16px;
            cursor:pointer;
        }
        .error{
            background:#fee2e2;
            color:#991b1b;
            border:1px solid #fecaca;
            padding:10px;
            border-radius:12px;
            margin-bottom:12px;
            font-weight:bold;
        }
    </style>
</head>
<body>
    <div class="card">
        <div class="top">
            <h2>MBO DIARIO CHSG3</h2>
        </div>

        {% if error %}
            <div class="error">{{ error }}</div>
        {% endif %}

        <p>Ingresa la clave para acceder al registro MBO.</p>

        <form method="POST">
            <label>Clave de acceso</label>
            <input type="password" name="password" placeholder="Clave" autofocus required>
            <button type="submit">Ingresar</button>
        </form>
    </div>
</body>
</html>
"""


@app.before_request
def proteger_app():
    password = os.environ.get("APP_PASSWORD", "").strip()

    # Si APP_PASSWORD no está definido, no se exige login.
    # En nube se recomienda definir APP_PASSWORD en las variables de entorno.
    if not password:
        return None

    if request.endpoint in ("login",):
        return None

    if session.get("mbo_autenticado"):
        return None

    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    password = os.environ.get("APP_PASSWORD", "").strip()

    if not password:
        session["mbo_autenticado"] = True
        return redirect(url_for("index"))

    error = ""
    if request.method == "POST":
        recibido = request.form.get("password", "")
        if recibido == password:
            session["mbo_autenticado"] = True
            return redirect(url_for("index"))
        error = "Clave incorrecta."

    return render_template_string(HTML_LOGIN, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))



@app.route("/")
def index():
    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) AS total FROM items_mbo WHERE activo = 1")
    total_items = cur.fetchone()["total"]

    cur.execute("SELECT DISTINCT zona FROM items_mbo WHERE activo = 1 ORDER BY zona")
    zonas = [r["zona"] for r in cur.fetchall()]

    zona = request.args.get("zona") or (zonas[0] if zonas else "")

    cur.execute("""
        SELECT * FROM items_mbo
        WHERE activo = 1 AND zona = ?
        ORDER BY id
    """, (zona,))
    items_raw = cur.fetchall()

    items = []
    for item in items_raw:
        item_dict = dict(item)
        item_dict["valor_opciones"] = opciones_selector_valor(item_dict.get("referencia", ""))
        item_dict["convertir_horas"] = es_item_horas_operacion(item_dict)
        items.append(item_dict)

    niveles = []
    niveles_vistos = set()
    for item in items:
        nivel = limpiar_texto(item["nivel"])
        if nivel and nivel not in niveles_vistos:
            niveles.append(nivel)
            niveles_vistos.add(nivel)

    cur.execute("""
        SELECT usuario
        FROM registros_mbo
        WHERE usuario IS NOT NULL AND TRIM(usuario) <> ''
        ORDER BY id DESC
        LIMIT 1
    """)
    fila_operador = cur.fetchone()
    ultimo_operador = fila_operador["usuario"] if fila_operador else ""

    conn.close()

    ahora = datetime.now()
    return render_template_string(
        HTML_INDEX,
        zonas=zonas,
        zona=zona,
        niveles=niveles,
        items=items,
        total_items=total_items,
        fecha_hoy=ahora.strftime("%Y-%m-%d"),
        hora_actual=ahora.strftime("%H:%M"),
        ultimo_operador=ultimo_operador,
    )


@app.route("/guardar_zona", methods=["POST"])
def guardar_zona():
    ahora_guardado = datetime.now()
    fecha = request.form.get("fecha", "") or ahora_guardado.strftime("%Y-%m-%d")
    hora = request.form.get("hora", "") or ahora_guardado.strftime("%H:%M")
    usuario = limpiar_texto(request.form.get("usuario", ""))
    zona = request.form.get("zona", "")

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT * FROM items_mbo WHERE activo = 1 AND zona = ? ORDER BY id", (zona,))
    items = cur.fetchall()

    guardados = 0
    fecha_registro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for item in items:
        item_id = item["id"]
        valor = limpiar_texto(request.form.get(f"valor_{item_id}", ""))

        if valor and es_item_horas_operacion(item):
            valor_convertido = convertir_duracion_a_horas(valor)
            if valor_convertido is not None:
                valor = valor_convertido

        observacion = limpiar_texto(request.form.get(f"obs_{item_id}", ""))
        foto_file = request.files.get(f"foto_{item_id}")
        nombre_foto = ""

        if foto_file and foto_file.filename:
            ext = os.path.splitext(foto_file.filename)[1].lower()
            nombre_foto = secure_filename(f"MBO_{fecha}_{hora}_{zona}_{item_id}{ext}".replace(":", ""))
            foto_file.save(os.path.join(UPLOAD_DIR, nombre_foto))

        if not valor and not observacion and not nombre_foto:
            continue

        cur.execute("""
            INSERT INTO registros_mbo
            (fecha, hora, usuario, zona, item_id, nivel, sistema, equipo, descripcion,
             unidad, senal, seteos, referencia, valor, observacion, foto, fecha_registro, estado)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fecha,
            hora,
            usuario,
            zona,
            item_id,
            item["nivel"],
            item["sistema"],
            item["equipo"],
            item["descripcion"],
            item["unidad"],
            item["senal"],
            item["seteos"],
            item["referencia"],
            valor,
            observacion,
            nombre_foto,
            fecha_registro,
            "REGISTRADO",
        ))
        guardados += 1

    conn.commit()
    conn.close()

    flash(f"Se guardaron {guardados} registros para {zona}.")
    return redirect(url_for("index", zona=zona))


@app.route("/registros")
def registros():
    fecha = request.args.get("fecha", "")
    zona = request.args.get("zona", "")

    conn = conectar()
    cur = conn.cursor()

    cur.execute("SELECT DISTINCT zona FROM items_mbo ORDER BY zona")
    zonas = [r["zona"] for r in cur.fetchall()]

    sql = "SELECT * FROM registros_mbo WHERE 1=1"
    params = []
    if fecha:
        sql += " AND fecha = ?"
        params.append(fecha)
    if zona:
        sql += " AND zona = ?"
        params.append(zona)
    sql += " ORDER BY id DESC LIMIT 500"

    cur.execute(sql, params)
    registros = cur.fetchall()
    conn.close()

    return render_template_string(HTML_REGISTROS, registros=registros, zonas=zonas, fecha=fecha, zona=zona)


@app.route("/uploads/<filename>")
def ver_foto(filename):
    return send_from_directory(UPLOAD_DIR, filename)


def valor_celda_con_merge(ws, fila, col):
    """Devuelve el valor real aunque la celda esté dentro de un rango combinado."""
    celda = ws.cell(fila, col)
    if celda.value is not None:
        return celda.value

    coord = celda.coordinate
    for rango in ws.merged_cells.ranges:
        if coord in rango:
            return ws.cell(rango.min_row, rango.min_col).value
    return None


def convertir_fecha(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    texto = limpiar_texto(valor)
    if not texto:
        return None

    # Evita interpretar fórmulas como fechas.
    if texto.startswith("="):
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto[:10], fmt).date()
        except ValueError:
            pass
    return None


def convertir_hora(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.time().replace(second=0, microsecond=0)
    if isinstance(valor, time):
        return valor.replace(second=0, microsecond=0)
    if isinstance(valor, (int, float)):
        # Hora de Excel como fracción de día.
        segundos = int(round(float(valor) * 24 * 60 * 60)) % (24 * 60 * 60)
        return (datetime(1900, 1, 1) + timedelta(seconds=segundos)).time().replace(second=0, microsecond=0)

    texto = limpiar_texto(valor)
    if not texto or texto.startswith("="):
        return None

    texto = texto.replace(".", ":")
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(texto, fmt).time().replace(second=0, microsecond=0)
        except ValueError:
            pass
    return None


def copiar_formato_columna(ws, col_origen, col_destino):
    for fila in range(1, ws.max_row + 1):
        origen = ws.cell(fila, col_origen)
        destino = ws.cell(fila, col_destino)
        if origen.has_style:
            destino.font = copy(origen.font)
            destino.fill = copy(origen.fill)
            destino.border = copy(origen.border)
            destino.alignment = copy(origen.alignment)
            destino.number_format = origen.number_format
            destino.protection = copy(origen.protection)

    ws.column_dimensions[ws.cell(1, col_destino).column_letter].width = ws.column_dimensions[ws.cell(1, col_origen).column_letter].width


def buscar_o_crear_columna_registro(ws, fecha_txt, hora_txt):
    """
    Busca la columna de la fecha/hora del MBO.
    Si no existe, crea una columna nueva al final conservando el estilo de la última columna.
    """
    fecha_obj = convertir_fecha(fecha_txt)
    hora_obj = convertir_hora(hora_txt)

    cols = detectar_columnas(ws) or {}
    fila_header = cols.get("fila_header", 5)

    # En las hojas normales la fecha está en la fila de encabezado y la hora en la siguiente.
    for col in range(1, ws.max_column + 1):
        fecha_col = convertir_fecha(valor_celda_con_merge(ws, fila_header, col))
        hora_col = convertir_hora(valor_celda_con_merge(ws, fila_header + 1, col))
        if fecha_col == fecha_obj and (hora_obj is None or hora_col == hora_obj):
            return col

    # Si no encontró la hora exacta, acepta la fecha. Útil cuando la hoja no tiene hora visible.
    for col in range(1, ws.max_column + 1):
        fecha_col = convertir_fecha(valor_celda_con_merge(ws, fila_header, col))
        if fecha_col == fecha_obj:
            return col

    # Crear una columna nueva al final.
    col_nueva = ws.max_column + 1
    col_origen = max(col_nueva - 1, 1)
    copiar_formato_columna(ws, col_origen, col_nueva)

    celda_fecha = ws.cell(fila_header, col_nueva)
    celda_fecha.value = fecha_obj or fecha_txt
    celda_fecha.number_format = "dd-mmm"

    celda_hora = ws.cell(fila_header + 1, col_nueva)
    celda_hora.value = hora_obj or hora_txt
    celda_hora.number_format = "hh:mm"

    return col_nueva


def agregar_hoja_resumen(wb, registros):
    if "REGISTROS_MBO" in wb.sheetnames:
        del wb["REGISTROS_MBO"]

    ws = wb.create_sheet("REGISTROS_MBO")
    headers = [
        "ID", "FECHA", "HORA", "OPERADOR", "ZONA", "NIVEL", "SISTEMA", "EQUIPO",
        "DESCRIPCION", "UNIDAD", "SEÑAL", "SETEOS", "REFERENCIA", "VALOR",
        "OBSERVACION", "FOTO", "FECHA_REGISTRO", "ESTADO"
    ]
    ws.append(headers)

    for r in registros:
        ws.append([
            r["id"], r["fecha"], r["hora"], r["usuario"], r["zona"], r["nivel"],
            r["sistema"], r["equipo"], r["descripcion"], r["unidad"], r["senal"],
            r["seteos"], r["referencia"], r["valor"], r["observacion"], r["foto"],
            r["fecha_registro"], r["estado"]
        ])

    header_fill = PatternFill("solid", fgColor="005BAA")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    widths = {
        "A": 8, "B": 12, "C": 10, "D": 18, "E": 18, "F": 15, "G": 35, "H": 28,
        "I": 40, "J": 12, "K": 14, "L": 25, "M": 25, "N": 15, "O": 35,
        "P": 24, "Q": 20, "R": 15
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


@app.route("/exportar_excel")
def exportar_excel():
    fecha = request.args.get("fecha", "")
    zona = request.args.get("zona", "")

    ruta_excel = buscar_excel_mbo()
    if not ruta_excel:
        flash("No se encontró el Excel MBO base para llenar el formato.")
        return redirect(url_for("index"))

    conn = conectar()
    cur = conn.cursor()

    sql = """
        SELECT r.*, i.fila_excel
        FROM registros_mbo r
        LEFT JOIN items_mbo i ON r.item_id = i.id
        WHERE 1=1
    """
    params = []
    if fecha:
        sql += " AND r.fecha = ?"
        params.append(fecha)
    if zona:
        sql += " AND r.zona = ?"
        params.append(zona)
    sql += " ORDER BY r.fecha, r.hora, r.zona, r.id"

    cur.execute(sql, params)
    registros = cur.fetchall()
    conn.close()

    wb = load_workbook(ruta_excel)

    # Deja solo las hojas principales del MBO. Las hojas auxiliares tipo FOR se retiran del exportado.
    for nombre_hoja in list(wb.sheetnames):
        if nombre_hoja not in SHEETS_MBO:
            del wb[nombre_hoja]

    columnas_por_evento = {}

    for r in registros:
        # COLCHON DE AIRE: OPERADOR DE TURNO es parte del footer/firma del formato,
        # no un punto de inspección. No se escribe como dato.
        if es_registro_operador_turno_colchon(r):
            continue

        nombre_hoja = r["zona"]
        fila_excel = r["fila_excel"]
        valor = r["valor"]

        if not nombre_hoja or nombre_hoja not in wb.sheetnames:
            continue
        if not fila_excel:
            continue

        ws = wb[nombre_hoja]
        clave = (nombre_hoja, r["fecha"], r["hora"])

        if clave not in columnas_por_evento:
            columnas_por_evento[clave] = buscar_o_crear_columna_registro(ws, r["fecha"], r["hora"])

        col_valor = columnas_por_evento[clave]
        ws.cell(int(fila_excel), col_valor).value = valor
        ws.cell(int(fila_excel), col_valor).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    agregar_hoja_resumen(wb, registros)

    nombre = f"MBO_FORMATO_LLENADO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta = os.path.join(EXPORT_DIR, nombre)
    wb.save(ruta)

    return send_file(ruta, as_attachment=True)

@app.route("/recargar_items")
def recargar_items():
    total, mensaje = cargar_items_desde_excel(forzar=True)
    aplicar_ajustes_comunes()
    aplicar_ajustes_colchon_aire_solo_app()
    flash(f"{mensaje} Ítems cargados: {total}. Ajustes aplicados: COMUNES BARRA A/B y COLCHON DE AIRE sin mostrar OPERADOR DE TURNO como ítem.")
    return redirect(url_for("index"))


def inicializar_app():
    """
    Inicializa base de datos e ítems al arrancar.

    Importante:
    En Render/Gunicorn el bloque if __name__ == "__main__" NO se ejecuta.
    Por eso esta inicialización debe ejecutarse al importar app.py.
    """
    crear_bd()
    total, mensaje = cargar_items_desde_excel(forzar=False)
    aplicar_ajustes_comunes()
    aplicar_ajustes_colchon_aire_solo_app()
    print(mensaje)
    print(f"Ítems disponibles: {total}")


# Ejecutar inicialización también cuando la app corre con Gunicorn en Render.
inicializar_app()


if __name__ == "__main__":
    print("Abre desde esta PC: http://127.0.0.1:5000")
    print("Desde celular usa la IP que aparezca abajo, ejemplo: http://192.168.1.102:5000")

    puerto = int(os.environ.get("PORT", "5000"))

    app.run(
        host="0.0.0.0",
        port=puerto,
        debug=False,
        use_reloader=False
    )
